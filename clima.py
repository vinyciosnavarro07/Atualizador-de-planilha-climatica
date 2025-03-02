import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from datetime import datetime

def buscar_previsao():
    try:
        navegador = webdriver.Chrome()

        navegador.get('https://www.climatempo.com.br/previsao-do-tempo/cidade/558/saopaulo-sp')

        WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="max-temp-1"]'))
        )

        valor_temperatura = navegador.find_element(By.XPATH, '//*[@id="max-temp-1"]').text

        WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[2]/div[1]/ul/li[4]/div/p/span[1]'))
        )

        valor_umidade = navegador.find_element(By.XPATH, '//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[2]/div[1]/ul/li[4]/div/p/span[1]').text

        data_hora_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        caminho_arquivo = 'temperatura.xlsx'

        if os.path.exists(caminho_arquivo):
            arquivo = load_workbook(caminho_arquivo)
            planilha = arquivo.active
        else:
            arquivo = Workbook()
            planilha = arquivo.active
            planilha.title = 'Previsao Tempo'
            planilha['A1'] = 'Data/Hora'
            planilha['B1'] = 'Temperatura'
            planilha['C1'] = 'Umidade'

        linha_disponivel = planilha.max_row + 1

        planilha[f'A{linha_disponivel}'] = data_hora_atual
        planilha[f'B{linha_disponivel}'] = valor_temperatura
        planilha[f'C{linha_disponivel}'] = valor_umidade

        arquivo.save(caminho_arquivo)

        messagebox.showinfo("Sucesso", "Valores salvos com sucesso na planilha!")

        navegador.quit()

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

janela = tk.Tk()
janela.title("Atualizar Previsão do Tempo")

janela.geometry("300x150")

label = tk.Label(janela, text="Clique no botão para atualizar a previsão na planilha.")
label.pack(pady=20)

botao = tk.Button(janela, text="Buscar Previsão", command=buscar_previsao)
botao.pack()

janela.mainloop()